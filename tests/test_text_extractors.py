"""Test degli estrattori di testo per i formati a lettura diretta
(pipeline/text_extractors.py): .xml, .doc legacy, fogli di calcolo e Office/ODF."""

import tempfile
import unittest
import zipfile
from pathlib import Path

from pipeline.text_extractors import (
    extract_csv_text,
    extract_doc_text,
    extract_ods_text,
    extract_odp_text,
    extract_odt_text,
    extract_pptx_text,
    extract_xlsx_text,
    extract_xml_text,
)

_FIXTURE_DOC = Path(__file__).parent / "fixtures" / "sample_legacy.doc"

_ODS_CONTENT = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
 xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
 xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"
 xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0">
 <office:body><office:spreadsheet>
  <table:table table:name="Foglio1">
   <table:table-row>
    <table:table-cell><text:p>Nome</text:p></table:table-cell>
    <table:table-cell><text:p>Importo</text:p></table:table-cell>
   </table:table-row>
   <table:table-row>
    <table:table-cell><text:p>Mario</text:p></table:table-cell>
    <table:table-cell office:value-type="float" office:value="100">
     <text:p>100</text:p></table:table-cell>
   </table:table-row>
  </table:table>
 </office:spreadsheet></office:body>
</office:document-content>"""

_ODT_CONTENT = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
 xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
 xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0">
 <office:body><office:text>
  <text:h>Titolo Documento</text:h>
  <text:p>Primo paragrafo di prova.</text:p>
  <text:p>Secondo paragrafo.</text:p>
 </office:text></office:body>
</office:document-content>"""

_ODP_CONTENT = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
 xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
 xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"
 xmlns:draw="urn:oasis:names:tc:opendocument:xmlns:drawing:1.0">
 <office:body><office:presentation>
  <draw:page draw:name="p1"><draw:frame><draw:text-box>
   <text:p>Prima diapositiva</text:p>
  </draw:text-box></draw:frame></draw:page>
  <draw:page draw:name="p2"><draw:frame><draw:text-box>
   <text:p>Seconda diapositiva</text:p>
  </draw:text-box></draw:frame></draw:page>
 </office:presentation></office:body>
</office:document-content>"""


def _make_zip(path: Path, members: dict) -> None:
    with zipfile.ZipFile(path, "w") as zf:
        for name, content in members.items():
            zf.writestr(name, content)


def _pptx_slide(text: str) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
        "<p:cSld><p:spTree><p:sp><p:txBody>"
        f"<a:p><a:r><a:t>{text}</a:t></a:r></a:p>"
        "</p:txBody></p:sp></p:spTree></p:cSld></p:sld>"
    )


class ExtractXmlTests(unittest.TestCase):
    def test_extracts_node_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.xml"
            p.write_text(
                "<?xml version='1.0'?>\n<root><a>Alfa</a><b>Beta</b></root>",
                encoding="utf-8",
            )
            text = extract_xml_text(p)
        self.assertIn("Alfa", text)
        self.assertIn("Beta", text)

    def test_malformed_xml_falls_back_to_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "rotto.xml"
            # XML non ben formato: il fallback lenient/raw deve comunque dare testo.
            p.write_text("<root><a>Contenuto</b></root", encoding="utf-8")
            text = extract_xml_text(p)
        self.assertIn("Contenuto", text)

    def test_namespaced_xml(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "ns.xml"
            p.write_text(
                "<n:doc xmlns:n='urn:x'><n:t>Testo con namespace</n:t></n:doc>",
                encoding="utf-8",
            )
            text = extract_xml_text(p)
        self.assertIn("Testo con namespace", text)


class ExtractDocLegacyTests(unittest.TestCase):
    def test_non_ole_raises_clear_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "finto.doc"
            p.write_bytes(b"questo non e' un file OLE2")
            with self.assertRaises(RuntimeError):
                extract_doc_text(p)

    def test_empty_file_raises_runtime_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "vuoto.doc"
            p.write_bytes(b"")
            with self.assertRaises(RuntimeError):
                extract_doc_text(p)

    @unittest.skipUnless(_FIXTURE_DOC.exists(), "fixture assente")
    def test_truncated_ole_raises_runtime_not_raw(self):
        # Un .doc OLE valido ma troncato non deve far emergere struct.error/OLE
        # grezza: il contratto è un RuntimeError con messaggio guida.
        raw = _FIXTURE_DOC.read_bytes()
        for cut in (256, 512, 1024):
            with tempfile.TemporaryDirectory() as tmp:
                p = Path(tmp) / "tronco.doc"
                p.write_bytes(raw[:cut])
                with self.assertRaises(RuntimeError):
                    extract_doc_text(p)

    @unittest.skipUnless(
        _FIXTURE_DOC.exists(),
        "fixture sample_legacy.doc assente",
    )
    def test_extracts_real_binary_doc(self):
        # Fixture: testWORD.doc di Apache Tika (Apache-2.0), un vero Word 97-2003.
        text = extract_doc_text(_FIXTURE_DOC)
        self.assertIn("Sample Word Document", text)
        self.assertIn("Heading Level 1", text)
        # Il testo delle celle di tabella viene estratto (flatten, come per .docx).
        self.assertIn("nested table", text.lower())
        # Nessun marcatore di campo grezzo (0x13/0x14/0x15) nell'output.
        self.assertNotIn("\x13", text)
        self.assertNotIn("\x14", text)
        self.assertNotIn("\x15", text)


class ExtractXlsxTests(unittest.TestCase):
    def test_cells_to_markdown_table(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Foglio1"
            ws.append(["Nome", "Importo"])
            ws.append(["Mario", 100])
            ws.append(["Luigi", 250])
            wb.save(p)
            text = extract_xlsx_text(p)
        self.assertIn("Nome", text)
        self.assertIn("Importo", text)
        self.assertIn("Mario", text)
        self.assertIn("100", text)
        self.assertIn("---", text)  # separatore di tabella Markdown
        self.assertNotIn("100.0", text)  # interi senza coda ".0"

    def test_multiple_sheets_get_name_headers(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "multi.xlsx"
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Primo"
            ws1.append(["A"])
            ws1.append(["x"])
            ws2 = wb.create_sheet("Secondo")
            ws2.append(["B"])
            ws2.append(["y"])
            wb.save(p)
            text = extract_xlsx_text(p)
        self.assertIn("## Primo", text)
        self.assertIn("## Secondo", text)

    def test_dates_formatted_not_serial(self):
        import datetime

        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "date.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Scadenza"])
            ws.append([datetime.date(2021, 1, 1)])
            wb.save(p)
            text = extract_xlsx_text(p)
        self.assertIn("01/01/2021", text)  # data formattata, non "44197"
        self.assertNotIn("44197", text)


class ExtractCsvTests(unittest.TestCase):
    def test_comma_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.csv"
            p.write_text("Nome,Citta\nMario,Roma\n", encoding="utf-8")
            text = extract_csv_text(p)
        self.assertIn("| Nome | Citta |", text)
        self.assertIn("| Mario | Roma |", text)

    def test_semicolon_delimiter_detected(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.csv"
            p.write_text("a;b;c\n1;2;3\n", encoding="utf-8")
            text = extract_csv_text(p)
        self.assertIn("| a | b | c |", text)
        self.assertIn("| 1 | 2 | 3 |", text)

    def test_tsv_uses_tab(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.tsv"
            p.write_text("x\ty\n10\t20\n", encoding="utf-8")
            text = extract_csv_text(p)
        self.assertIn("| x | y |", text)
        self.assertIn("| 10 | 20 |", text)


class ExtractOdsTests(unittest.TestCase):
    def test_table_to_markdown(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "dati.ods"
            _make_zip(p, {"content.xml": _ODS_CONTENT})
            text = extract_ods_text(p)
        self.assertIn("Nome", text)
        self.assertIn("Importo", text)
        self.assertIn("Mario", text)
        self.assertIn("100", text)
        self.assertIn("---", text)


class ExtractOdtTests(unittest.TestCase):
    def test_paragraphs_and_heading(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "doc.odt"
            _make_zip(p, {"content.xml": _ODT_CONTENT})
            text = extract_odt_text(p)
        self.assertIn("Titolo Documento", text)
        self.assertIn("Primo paragrafo di prova.", text)
        self.assertIn("Secondo paragrafo.", text)


class ExtractOdpTests(unittest.TestCase):
    def test_slides_with_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "pres.odp"
            _make_zip(p, {"content.xml": _ODP_CONTENT})
            text = extract_odp_text(p)
        self.assertIn("## Diapositiva 1", text)
        self.assertIn("## Diapositiva 2", text)
        self.assertIn("Prima diapositiva", text)
        self.assertIn("Seconda diapositiva", text)


class ExtractPptxTests(unittest.TestCase):
    def test_slides_in_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "pres.pptx"
            _make_zip(p, {
                "ppt/slides/slide1.xml": _pptx_slide("Titolo PPTX"),
                "ppt/slides/slide2.xml": _pptx_slide("Contenuto due"),
                "ppt/presentation.xml": "<x/>",
            })
            text = extract_pptx_text(p)
        self.assertIn("## Diapositiva 1", text)
        self.assertIn("## Diapositiva 2", text)
        self.assertIn("Titolo PPTX", text)
        self.assertIn("Contenuto due", text)
        self.assertLess(text.index("Titolo PPTX"), text.index("Contenuto due"))


if __name__ == "__main__":
    unittest.main()
