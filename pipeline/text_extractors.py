"""Plain-text extraction for document formats that need no OCR."""

import datetime
import re
from pathlib import Path


def extract_docx_text(file_path: Path) -> str:
    import docx

    doc = docx.Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def extract_html_text(file_path: Path) -> str:
    from bs4 import BeautifulSoup

    html_content = file_path.read_text(encoding="utf-8", errors="replace")
    parts: list[str] = []

    saved_url_match = re.search(
        r'<!--\s*saved from url=\(\d+\)(.*?)\s*-->',
        html_content,
    )
    if saved_url_match:
        parts.append(f"Source URL: {saved_url_match.group(1).strip()}")

    soup = BeautifulSoup(html_content, "html.parser")

    if soup.title and soup.title.string:
        parts.append(f"Title: {soup.title.string.strip()}")

    for meta_name in ("description", "author", "keywords"):
        meta_tag = soup.find("meta", attrs={"name": lambda x: x and x.lower() == meta_name})
        if meta_tag and meta_tag.get("content"):
            parts.append(f"{meta_name.capitalize()}: {meta_tag.get('content').strip()}")

    canonical = soup.find("link", rel="canonical")
    if canonical and canonical.get("href") and not saved_url_match:
        parts.append(f"Canonical URL: {canonical.get('href').strip()}")

    if parts:
        parts.append("\n--- TESTO DEL SITO ---")

    parts.append(soup.get_text(separator="\n", strip=True))
    return "\n".join(parts)


def extract_rtf_text(file_path: Path) -> str:
    from striprtf.striprtf import rtf_to_text

    rtf_content = file_path.read_text(encoding="utf-8", errors="replace")
    return rtf_to_text(rtf_content)


def extract_xml_text(file_path: Path) -> str:
    """Estrae il testo leggibile da un file XML.

    Prova il parser XML della stdlib (ElementTree, testo dei nodi in ordine di
    documento); in fallback usa BeautifulSoup in modalita' lenient; ultimo
    fallback: contenuto grezzo. Nessuna dipendenza nuova.
    """
    raw = file_path.read_text(encoding="utf-8", errors="replace")

    try:
        import xml.etree.ElementTree as ET

        root = ET.fromstring(raw)
        parts = [t.strip() for t in root.itertext() if t and t.strip()]
        text = "\n".join(parts)
        if text.strip():
            return text
    except Exception:
        pass

    try:
        from bs4 import BeautifulSoup

        text = BeautifulSoup(raw, "html.parser").get_text("\n", strip=True)
        if text.strip():
            return text
    except Exception:
        pass

    return raw


def extract_doc_text(file_path: Path) -> str:
    """Estrae il testo da un file Word 97-2003 binario (.doc) legacy.

    python-docx apre solo il formato OOXML (.docx), non il vecchio .doc binario
    OLE2. Qui leggiamo direttamente lo stream ``WordDocument`` + la piece table
    (CLX/Pcdt) tramite ``olefile`` — gia' presente nel bundle come dipendenza di
    ``extract-msg``, pure-python e senza richiedere Word o LibreOffice. Se il file
    e' cifrato, danneggiato o in un layout non gestito, solleviamo un errore
    chiaro che invita a risalvarlo come .docx/.rtf.
    """
    import struct

    import olefile

    def _clean(text: str) -> str:
        for src, dst in (
            ("\r", "\n"), ("\x07", "\n"), ("\x0b", "\n"), ("\x0c", "\n"),
            ("\x1e", "-"), ("\xa0", " "),
            ("\x13", ""), ("\x14", ""), ("\x15", ""),
        ):
            text = text.replace(src, dst)
        text = re.sub(r'\s*HYPERLINK\s+("[^"]*"|\\l\s+"[^"]*"|\\h)\s*', " ", text)
        text = "".join(c for c in text if c >= " " or c in "\t\n")
        return re.sub(r"[ \t]{2,}", " ", text).strip()

    if not olefile.isOleFile(str(file_path)):
        raise RuntimeError(
            "Il file .doc non e' un documento Word 97-2003 valido (OLE2)."
        )

    # Contratto: qualunque .doc non interpretabile (OLE corrotto, FIB/CLX
    # troncata, stream mancanti) deve diventare un RuntimeError con messaggio
    # guida, mai un'eccezione grezza (struct.error/OLE/IndexError). I RuntimeError
    # con messaggi specifici sollevati qui dentro vengono ri-propagati invariati.
    try:
        ole = olefile.OleFileIO(str(file_path))
    except Exception as e:
        raise RuntimeError(f"file .doc non apribile (OLE non valido): {e}") from e

    try:
        if ole.exists("EncryptedPackage") or ole.exists("Encryption"):
            raise RuntimeError("documento .doc cifrato")
        if not ole.exists("WordDocument"):
            raise RuntimeError("stream WordDocument assente")

        wd = ole.openstream("WordDocument").read()
        if wd[:2] != b"\xec\xa5":  # magic 0xA5EC della FIB
            raise RuntimeError("intestazione FIB non valida")
        if len(wd) < 0x0200:  # la FIB deve coprire gli offset fissi sotto
            raise RuntimeError("FIB del .doc troncata o non valida")

        flags = struct.unpack_from("<H", wd, 0x000A)[0]
        if flags & 0x0100:  # fEncrypted
            raise RuntimeError("documento .doc cifrato")
        fc_min = struct.unpack_from("<i", wd, 0x0018)[0]
        fc_mac = struct.unpack_from("<i", wd, 0x001C)[0]
        table_name = "1Table" if (flags & 0x0200) else "0Table"

        text = ""
        if ole.exists(table_name):
            tbl = ole.openstream(table_name).read()
            fc_clx = struct.unpack_from("<i", wd, 0x01A2)[0]
            lcb_clx = struct.unpack_from("<i", wd, 0x01A6)[0]
            clx = tbl[fc_clx:fc_clx + lcb_clx]

            pcdt = None
            i = 0
            while i < len(clx):
                if clx[i] == 0x02:  # blocco Pcdt (piece table)
                    cb = struct.unpack_from("<I", clx, i + 1)[0]
                    pcdt = clx[i + 5:i + 5 + cb]
                    break
                if clx[i] == 0x01:  # Prc da saltare
                    cb = struct.unpack_from("<H", clx, i + 1)[0]
                    i += 3 + cb
                else:
                    break

            if pcdt:
                n = (len(pcdt) - 4) // 12
                cps = [
                    struct.unpack_from("<I", pcdt, k * 4)[0]
                    for k in range(n + 1)
                ]
                base = (n + 1) * 4
                chunks = []
                for k in range(n):
                    fcf = struct.unpack_from("<I", pcdt, base + k * 8 + 2)[0]
                    compressed = bool(fcf & 0x40000000)
                    fc = fcf & 0x3FFFFFFF
                    length = cps[k + 1] - cps[k]
                    if compressed:  # cp1252, offset dimezzato
                        chunks.append(
                            wd[fc // 2: fc // 2 + length].decode("cp1252", "replace")
                        )
                    else:  # UTF-16LE
                        chunks.append(
                            wd[fc: fc + length * 2].decode("utf-16-le", "replace")
                        )
                text = "".join(chunks)

        if not text and 0 <= fc_min < fc_mac <= len(wd):
            # Documento non "complex" (nessuna piece table): testo contiguo cp1252.
            text = wd[fc_min:fc_mac].decode("cp1252", "replace")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"struttura .doc non interpretabile: {e}") from e
    finally:
        ole.close()

    cleaned = _clean(text)
    # Se quasi tutto e' illeggibile, meglio un errore chiaro che testo-spazzatura.
    if not cleaned or cleaned.count("�") > len(cleaned) // 3:
        raise RuntimeError(
            "Impossibile estrarre testo leggibile dal .doc legacy (forse cifrato, "
            "danneggiato o in un formato non gestito). Salvalo come .docx o .rtf e "
            "riprova."
        )
    return cleaned


# ── Helper condivisi per fogli di calcolo e formati ZIP+XML (ODF/OOXML) ──────

# Le ripetizioni di celle/righe vuote in ODS (es. number-columns-repeated="1024")
# vengono limitate per non generare tabelle enormi: tanto le code vuote sono poi
# rifilate da _rows_to_markdown.
_ODF_MAX_REPEAT = 100


def _safe_int(value, default: int = 1) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _local_name(tag: str) -> str:
    """Nome locale di un tag/attributo ElementTree, senza prefisso namespace."""
    return tag.rsplit("}", 1)[-1]


def _attr_by_local(el, local: str) -> str | None:
    """Valore di un attributo cercato per nome locale (namespace-agnostico)."""
    for key, value in el.attrib.items():
        if _local_name(key) == local:
            return value
    return None


def _fmt_cell(value) -> str:
    """Converte un valore di cella (es. da openpyxl) in stringa leggibile."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, datetime.datetime):
        if value.hour or value.minute or value.second:
            base = value.strftime("%d/%m/%Y %H:%M")
            return f"{base}:{value.second:02d}" if value.second else base
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _rows_to_markdown(rows: list[list]) -> str:
    """Rende una griglia di celle come tabella Markdown (1a riga = intestazione).

    Rifila le righe/colonne vuote ai bordi ed esegue l'escape dei caratteri ``|``.
    Restituisce stringa vuota se non resta alcuna cella con contenuto.
    """
    grid = [[("" if c is None else str(c)) for c in row] for row in rows]

    def row_empty(r: list[str]) -> bool:
        return not any(c.strip() for c in r)

    while grid and row_empty(grid[0]):
        grid.pop(0)
    while grid and row_empty(grid[-1]):
        grid.pop()
    if not grid:
        return ""

    width = max(len(r) for r in grid)

    def col_empty(idx: int) -> bool:
        return all(idx >= len(r) or not r[idx].strip() for r in grid)

    while width > 0 and col_empty(width - 1):
        width -= 1
    if width == 0:
        return ""

    def cell(r: list[str], idx: int) -> str:
        text = r[idx] if idx < len(r) else ""
        return text.replace("|", "\\|").replace("\n", " ").strip()

    header = "| " + " | ".join(cell(grid[0], i) for i in range(width)) + " |"
    separator = "| " + " | ".join("---" for _ in range(width)) + " |"
    body = [
        "| " + " | ".join(cell(r, i) for i in range(width)) + " |"
        for r in grid[1:]
    ]
    return "\n".join([header, separator, *body])


def extract_xlsx_text(file_path: Path) -> str:
    """Estrae il testo da un foglio di calcolo Excel (.xlsx/.xlsm).

    Usa openpyxl in sola lettura coi valori calcolati (``data_only``): ogni foglio
    diventa una tabella Markdown, preceduta dal nome del foglio quando il file ne
    contiene piu' d'uno. openpyxl (pure-python) gestisce date e formati numerici,
    che un parsing manuale dell'XML perderebbe.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheets = wb.worksheets
        parts: list[str] = []
        for ws in sheets:
            rows = [
                [_fmt_cell(v) for v in row]
                for row in ws.iter_rows(values_only=True)
            ]
            table = _rows_to_markdown(rows)
            if not table:
                continue
            parts.append(f"## {ws.title}\n\n{table}" if len(sheets) > 1 else table)
        return "\n\n".join(parts)
    finally:
        wb.close()


def extract_csv_text(file_path: Path) -> str:
    """Estrae una tabella Markdown da un file CSV/TSV.

    Rileva il delimitatore (``,`` ``;`` tab ``|``); per i ``.tsv`` usa sempre il tab.
    """
    import csv
    import io

    raw = file_path.read_text(encoding="utf-8-sig", errors="replace")
    if file_path.suffix.lower() == ".tsv":
        delimiter = "\t"
    else:
        try:
            delimiter = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            sample = raw[:4096]
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
    rows = list(csv.reader(io.StringIO(raw), delimiter=delimiter))
    return _rows_to_markdown(rows)


def _odf_content_root(file_path: Path):
    """Radice ElementTree di ``content.xml`` dentro un contenitore ODF (ZIP)."""
    import xml.etree.ElementTree as ET
    import zipfile

    with zipfile.ZipFile(file_path) as zf:
        raw = zf.read("content.xml")
    return ET.fromstring(raw)


def _odf_cell_text(cell) -> str:
    """Testo di una cella ODF: tutti i paragrafi/titoli concatenati per riga."""
    paragraphs: list[str] = []
    for node in cell.iter():
        if _local_name(node.tag) in ("p", "h"):
            text = "".join(node.itertext()).strip()
            if text:
                paragraphs.append(text)
    return "\n".join(paragraphs)


def _odf_table_rows(table) -> list[list[str]]:
    """Righe/celle di una ``table:table`` ODF, espandendo le ripetizioni (capate)."""
    rows: list[list[str]] = []
    for row_el in table:
        if _local_name(row_el.tag) != "table-row":
            continue
        rep_rows = min(
            _safe_int(_attr_by_local(row_el, "number-rows-repeated")), _ODF_MAX_REPEAT
        )
        cells: list[str] = []
        for cell_el in row_el:
            lname = _local_name(cell_el.tag)
            if lname not in ("table-cell", "covered-table-cell"):
                continue
            rep_cols = min(
                _safe_int(_attr_by_local(cell_el, "number-columns-repeated")),
                _ODF_MAX_REPEAT,
            )
            text = _odf_cell_text(cell_el) if lname == "table-cell" else ""
            cells.extend([text] * rep_cols)
        for _ in range(rep_rows):
            rows.append(list(cells))
    return rows


def extract_ods_text(file_path: Path) -> str:
    """Estrae il testo da un foglio di calcolo OpenDocument (.ods). Pure-python.

    Legge ``content.xml`` e rende ogni tabella come tabella Markdown, preceduta dal
    nome del foglio quando ce n'e' piu' d'uno.
    """
    root = _odf_content_root(file_path)
    rendered: list[tuple[str, str]] = []
    for table in root.iter():
        if _local_name(table.tag) != "table":
            continue
        md = _rows_to_markdown(_odf_table_rows(table))
        if md:
            rendered.append((_attr_by_local(table, "name") or "", md))
    if not rendered:
        return ""
    if len(rendered) > 1:
        return "\n\n".join(
            f"## {name}\n\n{md}" if name else md for name, md in rendered
        )
    return rendered[0][1]


def _render_odf_blocks(el, out: list[str], slide_counter: list[int]) -> None:
    """Visita in ordine di documento il corpo ODF, raccogliendo blocchi di testo."""
    lname = _local_name(el.tag)
    if lname == "page":  # draw:page → diapositiva di una presentazione (.odp)
        slide_counter[0] += 1
        out.append(f"## Diapositiva {slide_counter[0]}")
        for child in el:
            _render_odf_blocks(child, out, slide_counter)
        return
    if lname in ("p", "h"):  # paragrafo/titolo: itertext copre gia' span e link
        text = "".join(el.itertext()).strip()
        if text:
            out.append(text)
        return
    if lname == "table":  # tabella incorporata → tabella Markdown
        md = _rows_to_markdown(_odf_table_rows(el))
        if md:
            out.append(md)
        return
    for child in el:
        _render_odf_blocks(child, out, slide_counter)


def _extract_odf_document(file_path: Path) -> str:
    root = _odf_content_root(file_path)
    body = next((el for el in root.iter() if _local_name(el.tag) == "body"), root)
    blocks: list[str] = []
    _render_odf_blocks(body, blocks, [0])
    return "\n\n".join(blocks)


def extract_odt_text(file_path: Path) -> str:
    """Estrae il testo da un documento OpenDocument Writer (.odt). Pure-python."""
    return _extract_odf_document(file_path)


def extract_odp_text(file_path: Path) -> str:
    """Estrae il testo da una presentazione OpenDocument Impress (.odp).

    Ogni diapositiva (``draw:page``) e' introdotta da ``## Diapositiva N``. Pure-python.
    """
    return _extract_odf_document(file_path)


def extract_pptx_text(file_path: Path) -> str:
    """Estrae il testo da una presentazione PowerPoint (.pptx). Pure-python.

    Legge le slide ``ppt/slides/slideN.xml`` in ordine numerico e raccoglie il testo
    dei run ``a:t`` (una riga per paragrafo ``a:p``); ogni diapositiva e' introdotta
    da ``## Diapositiva N``.
    """
    import xml.etree.ElementTree as ET
    import zipfile

    slide_re = re.compile(r"ppt/slides/slide(\d+)\.xml$")
    parts: list[str] = []
    with zipfile.ZipFile(file_path) as zf:
        names = sorted(
            (n for n in zf.namelist() if slide_re.match(n)),
            key=lambda n: int(slide_re.match(n).group(1)),
        )
        for idx, name in enumerate(names, 1):
            root = ET.fromstring(zf.read(name))
            lines: list[str] = []
            for para in root.iter():
                if _local_name(para.tag) != "p":
                    continue
                runs = [
                    "".join(t.itertext())
                    for t in para.iter()
                    if _local_name(t.tag) == "t"
                ]
                line = "".join(runs).strip()
                if line:
                    lines.append(line)
            if lines:
                parts.append(f"## Diapositiva {idx}\n\n" + "\n".join(lines))
    return "\n\n".join(parts)
